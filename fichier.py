from datetime import timedelta
import hashlib
from openpyxl.styles.alignment import Alignment
from flask import Flask, jsonify, render_template, request
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils import range_boundaries
import dropbox
import pymysql
from flask import redirect
from openpyxl.utils import get_column_letter

from sqlalchemy.exc import SQLAlchemyError
from flask_sqlalchemy import SQLAlchemy
import mysql.connector
from flask_jwt_extended import (
    JWTManager, jwt_required, create_access_token,
    get_jwt_identity
)
from flask_cors import CORS

app = Flask(__name__)
CORS(app)
dbx = dropbox.Dropbox('sl.Be9exiJVom_SEuXquBL5a4EJ5Iuvc5ggxlj8rUKcVURskYW8ynst0-99_rPj1yfDGTgSlOC92KVULm0OVUSrIaHkrnQzRcWoeZNrIRdV2d17Fv5xRHQhFnaJLGDCF8djaCJjSGMO')
app.config['JWT_SECRET_KEY'] = "2D4A614E645267556B58703273357638782F413F4428472B4B6250655368566D"
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:@localhost/confidenzbd'
jwt = JWTManager(app)

db = SQLAlchemy(app)

db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'confidenzbd'
}

# Clé secrète en tant que string

class Entreprise(db.Model):
    id_entreprise = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(50))
    nomuser = db.Column(db.String(50),unique=True, nullable=False)
    telephone = db.Column(db.String(15), unique=True, nullable=False)
    email = db.Column(db.String(50), unique=True, nullable=False)
    mot_de_passe = db.Column(db.String(100))
    role = db.Column(db.Integer)
    
class Employe(db.Model):
    __tablename__ = 'employe'

    id_employe = db.Column(db.String(255), primary_key=True)
    nom = db.Column(db.String(255), nullable=False)
    prenom = db.Column(db.String(255), nullable=False)
    email = db.Column(db.String(255), nullable=False)
    mdp = db.Column(db.String(255), nullable=False)
    telephone = db.Column(db.BigInteger, nullable=False)
    poste = db.Column(db.String(255), nullable=False)
    role = db.Column(db.Integer, nullable=False)
    id_entreprise = db.Column(db.Integer, db.ForeignKey('entreprise.id_entreprise'), nullable=False)

    entreprise = db.relationship('Entreprise', backref='employes')
    
    

def get_database_connection():
    connection = mysql.connector.connect(**db_config)
    return connection

@app.route('/exist', methods=['POST'])
def isUserExist(nomuser):
    existing_entreprise = Entreprise.query.filter_by(nomuser = nomuser).first()
    if existing_entreprise:
        return jsonify(message='L\'utilisateur existe déjà'), 409
    existing_employe = Employe.query.filter_by(nomuser =  nomuser)
    if existing_employe:
        return jsonify(message='L\'utilisateur existe déjà'), 409

# Route pour l'inscription
@app.route('/register', methods=['POST'])
def register():
    try:
        # Récupérer les informations d'inscription depuis la requête
        nom = request.form.get('nom')
        nomuser = request.form.get('nomuser')
        telephone = request.form.get('telephone')
        email = request.form.get('email')
        mdp = request.form.get('mdp')
        role = request.form.get('Role')

        # Début de la transaction
        with db.session.begin_nested():
            # Vérifier si l'utilisateur existe déjà
            existing_user = Entreprise.query.filter_by(nomuser=nomuser).first()
            if existing_user:
                return jsonify(message='L\'utilisateur existe déjà'), 409

            # Créer un nouvel utilisateur
            new_user = Entreprise(
                nom=nom,
                nomuser=nomuser,
                telephone=telephone,
                email=email,
                mot_de_passe=mdp,
                role= role
            )

            # Ajouter le nouvel utilisateur à la base de données
            db.session.add(new_user)


        # Inclure le type de compte de l'utilisateur dans le payload du token
        userlog = {
            'nomuser': nomuser,
            'role': role
        }
        expiresToken= timedelta(days=1)
        token = create_access_token(identity=userlog, expires_delta=expiresToken)
    
        # Valider la transaction principale
        db.session.commit()
        
        # Renvoyer le token JWT dans la réponse
        return jsonify({'access_token': token}), 200

    except SQLAlchemyError as e:
        db.session.rollback()  # Annuler la transaction en cas d'erreur
        return jsonify(message='Erreur lors de l\'inscription'), 500
    
    

# Route pour l'authentification
@app.route('/loginentreprise', methods=['POST'])
def loginEntrepise():
    # Récupérer les informations de connexion depuis la requête
    nomuser = request.form.get('nomuser')
    mdp = request.form.get('mdp')
    
    existing_entreprise = Entreprise.query.filter_by(nomuser = nomuser, mot_de_passe = mdp).first()
    if existing_entreprise is None:
        return jsonify(message='utilisateur non existant'), 409
    else:
        role = existing_entreprise.role
        
    info = {
        'nomuser': nomuser,
        'role': role
    }
    
    dateExpire = timedelta(days= 1)
    
    token = create_access_token(identity= info , expires_delta= dateExpire)

    # Renvoyer le token JWT dans la réponse
    return jsonify({'access_token': token}), 200

# Route pour l'authentification
@app.route('/loginemploye', methods=['POST'])
def loginEmploye():
    # Récupérer les informations de connexion depuis la requête
    nomuser = request.form.get('nomuser')
    mdp = request.form.get('mdp')
    
    existing_employe = Employe.query.filter_by(id_employe =  nomuser, mdp = mdp).first()
    if existing_employe is None:
        return jsonify(message='utilisateur non existant'), 409
    else:
        role = existing_employe.role
        
    info = {
        'nomuser': nomuser,
        'role': role
    }
    
    dateExpire = timedelta(days= 1)
    
    token = create_access_token(identity= info , expires_delta= dateExpire)

    # Renvoyer le token JWT dans la réponse
    return jsonify({'access_token': token}), 200





















#
#création d'un dossier lors de l'incription
#
@app.route('/accueil/<string:folder>', methods=['GET'])
def home(folder):
    base_path = '/Dropbox/DataConfidenz/Entreprise'
    folder_path = base_path + '/' + folder

    try:
        dbx.files_get_metadata(folder_path)
        message = f"Le dossier '{folder}' existe déjà dans votre dépôt Dropbox."
    except dropbox.exceptions.ApiError as e:
        if e.error.is_path() and isinstance(e.error.get_path(), dropbox.files.LookupError):
            dbx.files_create_folder_v2(folder_path)
            message = f"Le dossier '{folder}' a été créé avec succès dans votre dépôt Dropbox."
        else:
            message = "Une erreur s'est produite lors de la vérification du dossier dans Dropbox."
  
    return message

#
#uploader un fichier
#
@app.route('/upload/<string:folder>', methods=['POST'])
def upload(folder):
    file = request.files['Fichier']
    
    # Chemin du dossier dans Dropbox
    folder_path = '/Dropbox/DataConfidenz/Entreprise/' + folder
    
    # Envoi du fichier dans le dossier correspondant
    try:
        dbx.files_upload(file.read(), folder_path + '/' + file.filename)
        message = f"Le fichier '{file.filename}' a été téléchargé avec succès dans le dossier '{folder}' de votre dépôt Dropbox."
    except dropbox.exceptions.ApiError:
        message = "Une erreur s'est produite lors du téléchargement du fichier dans Dropbox."

    return message

#
#supprimer un fichier
#
@app.route('/delete/<string:folder>/<string:filename>', methods=['DELETE'])
def delete(filename, folder):
    # Chemin complet du fichier dans Dropbox
    file_path = '/Dropbox/DataConfidenz/Entreprise/' + folder +'/'+ filename
    
    try:
        dbx.files_delete_v2(file_path)
        message = f"Le fichier '{filename}' a été supprimé avec succès de votre dépôt Dropbox."
    except dropbox.exceptions.ApiError:
        message = f"Une erreur s'est produite lors de la suppression du fichier '{filename}' dans Dropbox."

    return message

#
#mettre à jour un fichier
#
@app.route('/update/<string:folder>/<string:filename>', methods=['POST'])
def update(filename, folder):
    file = request.files['Fichier']
    
    # Chemin complet du fichier dans Dropbox
    file_path = '/Dropbox/DataConfidenz/Entreprise/' + folder +'/'+ filename
    
    try:
        dbx.files_upload(file.read(), file_path, mode=dropbox.files.WriteMode.overwrite)
        message = f"Le fichier '{filename}' a été mis à jour avec succès dans votre dépôt Dropbox."
    except dropbox.exceptions.ApiError:
        message = f"Une erreur s'est produite lors de la mise à jour du fichier '{filename}' dans Dropbox."

    return message

#
#liste de tous les fichiers
#
@app.route('/list/<string:folder>', methods=['GET'])
def list_files(folder):
    # Chemin complet du dossier dans Dropbox
    folder_path = '/Dropbox/DataConfidenz/Entreprise/' + folder
    
    try:
        result = dbx.files_list_folder(folder_path)
        file_list = [entry.name for entry in result.entries if isinstance(entry, dropbox.files.FileMetadata)]
        message = f"Liste des fichiers dans le dossier '{folder}': {file_list}"
    except dropbox.exceptions.ApiError:
        message = f"Une erreur s'est produite lors de la récupération de la liste des fichiers dans le dossier '{folder}'."

    return message

#
#générer un lien de partage
#
def generate_public_link(file_path):
    try:
        link = dbx.sharing_create_shared_link(file_path)
        return link.url
    except dropbox.exceptions.ApiError:
        return None

#
#visualiser le fichier
#
@app.route('/view/<string:folder>/<string:filename>', methods=['GET'])
def view_file(filename, folder):
    try:
        # Chemin complet du fichier dans Dropbox
        file_path = '/Dropbox/DataConfidenz/Entreprise/' + folder + '/' + filename

        # Construire le lien de prévisualisation Dropbox
        preview_link = f"https://www.dropbox.com/preview/{file_path}?role=personal"
        # preview_link = generate_public_link(file_path)
        # # Rediriger l'utilisateur vers le lien de     
        return jsonify({'preview_link': preview_link})
    except dropbox.exceptions.ApiError:
        return f"Une erreur s'est produite lors de la récupération du fichier '{filename}' depuis Dropbox."
    
#
#extraction de les informations:
#
@app.route('/extract', methods=['POST'])
def extract():
    # Récupérer le fichier Excel depuis la requête POST
    file = request.files['file']
    
    # Charger le fichier Excel avec openpyxl
    workbook = load_workbook(file)
    
    # Sélectionner la première feuille de calcul
    sheet = workbook.active
    
    # Créer une liste pour stocker les tableaux de chaque ligne
    all_rows = []
    header = []
    corps = []
    
    #traitement de fichier pour rétirer les informations
    lignes_fusionnees = []

    for groupe in sheet.merged_cells.ranges:
        for ligne in range(groupe.min_row, groupe.max_row + 1):
            if ligne not in lignes_fusionnees:
                lignes_fusionnees.append(ligne)
    compteur = 1
    
    #si le  fichier a trois entêtes
    if len(lignes_fusionnees) == 2:
       
        for row in sheet.iter_rows(values_only=True):
            if compteur <= len(lignes_fusionnees)+1:
                # Créer un nouveau tableau pour la ligne
                row_values = []
                # Parcourir chaque cellule de la ligne
                for cell in row:
                    if cell is not None:
                        row_values.append(cell)
                header.append(row_values)
                compteur = compteur + 1
            else:
                break
        # Ajouter le tableau de la ligne à la liste  
        all_rows.append(header)  
    
        # Parcourir chaque ligne du fichier Excel
        for ligne in range(len(lignes_fusionnees)+2, sheet.max_row + 1):
            
            # Créer un nouveau tableau pour la ligne
            row_values = []
            
            for colonne in range(1, sheet.max_column + 1):
                cellule = sheet.cell(row=ligne, column=colonne)
                valeur = cellule.value
                row_values.append(valeur)
            corps.append(row_values)        
        # Ajouter le tableau de la ligne à la liste
        all_rows.append(corps)
    
    #si le fichier a deux entêtes
    if len(lignes_fusionnees) == 1:
    
        for row in sheet.iter_rows(values_only=True):
            if compteur <= len(lignes_fusionnees) + 1:
                # Créer un nouveau tableau pour la ligne
                row_values = []
                # Parcourir chaque cellule de la ligne
                for cell in row:
                    if cell is not None:
                        row_values.append(cell)
                header.append(row_values)
                compteur = compteur + 1
            else:
                break
        # Ajouter le tableau de la ligne à la liste  
        all_rows.append(header)  
    
        # Parcourir chaque ligne du fichier Excel
        for ligne in range(len(lignes_fusionnees)+ 2, sheet.max_row+1):
            
            # Créer un nouveau tableau pour la ligne
            row_values = []
            
            for colonne in range(1, sheet.max_column + 1):
                cellule = sheet.cell(row=ligne, column=colonne)
                valeur = cellule.value
                row_values.append(valeur)
            corps.append(row_values)        
        # Ajouter le tableau de la ligne à la liste
        all_rows.append(corps)
        
        
    #si le fichier a une entête
    if len(lignes_fusionnees) == 0:
    
        for row in sheet.iter_rows(values_only=True):
            if compteur <= len(lignes_fusionnees) + 1:
                # Créer un nouveau tableau pour la ligne
                row_values = []
                # Parcourir chaque cellule de la ligne
                for cell in row:
                    if cell is not None:
                        row_values.append(cell)
                header.append(row_values)
                compteur = compteur + 1
            else:
                break
        # Ajouter le tableau de la ligne à la liste  
        all_rows.append(header)  
    
        # Parcourir chaque ligne du fichier Excel
        for ligne in range(len(lignes_fusionnees)+ 2, sheet.max_row+1):
            
            # Créer un nouveau tableau pour la ligne
            row_values = []
            
            for colonne in range(1, sheet.max_column + 1):
                cellule = sheet.cell(row=ligne, column=colonne)
                valeur = cellule.value
                row_values.append(valeur)
            corps.append(row_values)        
        # Ajouter le tableau de la ligne à la liste
        all_rows.append(corps)
            
    return all_rows

if __name__ == '__main__':
    app.run(debug=True)
